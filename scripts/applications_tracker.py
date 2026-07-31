#!/usr/bin/env python3
"""
applications_tracker.py — master Applications Tracker (xlsx) tool.

The xlsx (applications/APPLICATIONS_TRACKER.xlsx) is the SINGLE master view of
progress: one row per application, with clickable links to the original job
posting and to the local application.md (which links onward to cv.spec /
cover-letter / motivation-letter / outreach, etc.).

One row per application, built from the per-application record files.
(the 国外高校研究所 + 国外企业 sheets): title row, header row, then data,
unified into one sheet with 类型 (高校/研究所 vs 企业) + 岗位性质 (科研/工程)
columns plus progress-tracking columns (兴趣级 / 匹配度 / 进展记录 / 联系人).

Source of truth = each `applications/<slug>/application.md` (YAML frontmatter +
Block A table). The xlsx is GENERATED from those files (`build`), and edited
tracking fields can be written BACK to the frontmatter (`to-md`).

Visual encoding:
  * 状态 / 兴趣级 / 匹配度 cells are colour-coded (高/中/低 → green/amber/red).
  * 申请截止 turns orange/red when a deadline is near or overdue (and unsubmitted);
    跟进日期 turns gold when a follow-up is due. 岗位性质 is tinted 科研/工程.
  * Rows are banded in groups of 10 (subtle blue) for quick visual location.
Auto-sort (action view): ① unsubmitted with a near deadline, ② ready-to-submit
(by 兴趣级 then 匹配度), ③ submitted / in-flight, ④ closed — research roles win
ties over engineering (PhD-leverage). Re-sort freely via the auto-filter.

Usage:
  python3 scripts/applications_tracker.py build         # application.md -> xlsx (rebuild all rows)
  python3 scripts/applications_tracker.py list          # print a summary table to stdout
  python3 scripts/applications_tracker.py to-md [SLUG]  # xlsx -> application.md frontmatter (status/dates)
  python3 scripts/applications_tracker.py set SLUG KEY VALUE   # set one md frontmatter field, then rebuild
"""
import sys
import re
import shutil
import datetime
from pathlib import Path

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

REPO = Path(__file__).resolve().parents[1]
APPS_DIR = REPO / "applications"
XLSX = APPS_DIR / "APPLICATIONS_TRACKER.xlsx"
BACKUP_DIR = REPO / "outputs" / "tracker_backups"   # local + gitignored (outputs/); guards the working copy
SHEET = "Applications"
TODAY = datetime.date.today()

# (header_zh, record_key). Order = column order in the sheet.
COLUMNS = [
    ("编号", "id"),
    ("类型", "type"),
    ("岗位性质", "role_nature"),
    ("轴 / Axis", "axis"),
    ("名称", "name"),
    ("地点", "location"),
    ("职位", "position"),
    ("工作内容 / 课题", "role"),
    ("状态", "status"),
    ("主动联系人", "contact_flag"),
    ("兴趣级", "interest"),
    ("匹配度", "fit_display"),
    ("申请截止", "deadline"),
    ("时效提醒", "staleness"),
    ("投递日期", "applied_date"),
    ("开工日期", "start_date"),
    ("跟进日期", "follow_up_after"),
    ("最近动作", "last_action_date"),
    ("进展记录", "progress"),
    ("联系人", "contact_name"),
    ("联系 / 套磁情况", "outreach"),
    ("合同", "contract"),
    ("要求", "requirements"),
    ("联系人邮箱", "contact_email"),
    ("申请方式", "apply_route"),
    ("申请平台", "platform"),
    ("原职位链接", "url"),
    ("材料 / 往来记录", "materials"),
    ("其他 / 备注", "notes"),
]

# columns rendered as a clickable hyperlink (key -> cell label)
LINK_LABELS = {
    "url": "JD ↗",
}

# trackable fields that `to-md` writes back into the frontmatter
TRACK_TO_MD = ["status", "applied_date", "follow_up_after", "last_action_date", "deadline"]

# --- status semantics --------------------------------------------------------
# Full pipeline vocabulary (code -> 中文 / colour / stage-rank): pre-application
# stages, in-flight stages, positive outcomes, and closed outcomes.
STATUS_ZH = {
    "lead": "已发现", "researching": "调研中", "evaluating": "评估中",
    "evaluated": "已评估 · 待投", "preparing": "材料准备中",
    "drafted": "材料已起草 · 待投", "ready": "待投递",
    "in_progress": "进行中", "applied": "已投递", "acknowledged": "已确认收到",
    "screening": "初筛中", "oa": "笔试 / 测评", "tech-task": "笔试 / 任务",
    "interview": "面试中", "final": "终面 / onsite", "offer": "已录用",
    "negotiating": "议 offer", "accepted": "已接受",
    "rejected": "已拒", "declined": "已婉拒", "ghosted": "无回复",
    "withdrawn": "已撤回", "on_hold": "暂缓 / 冻结",
    "closed": "已关闭", "closed-before-application": "已关闭 · 未投",
}
STATUS_ZH_REV = {v: k for k, v in STATUS_ZH.items()}
STATUS_FILL = {
    "lead": "F2F2F2", "researching": "FBE5D6", "evaluating": "FFF2CC",
    "evaluated": "FFF2CC", "preparing": "FFE699",
    "drafted": "FFD966", "ready": "FFD966",
    "in_progress": "FCE4D6", "applied": "C6EFCE", "acknowledged": "A9D08E",
    "screening": "DDEBF7", "oa": "BDD7EE", "tech-task": "BDD7EE",
    "interview": "9DC3E6", "final": "8EAADB", "offer": "92D050",
    "negotiating": "C6E0B4", "accepted": "70AD47",
    "rejected": "FF0000", "declined": "FFC7CE", "ghosted": "D9D9D9",
    "withdrawn": "D9D9D9", "on_hold": "DBDBDB",
    "closed": "D9D9D9", "closed-before-application": "D9D9D9",
}
CLOSED_STATUS = {"rejected", "declined", "ghosted", "withdrawn", "on_hold", "closed", "closed-before-application"}
SUBMITTED_STATUS = {"applied", "acknowledged", "screening", "oa", "tech-task",
                    "interview", "final", "offer", "negotiating", "accepted"}
STAGE_RANK = {
    "accepted": 0, "offer": 1, "negotiating": 2, "final": 3, "interview": 4,
    "oa": 5, "tech-task": 5, "screening": 6, "acknowledged": 7, "applied": 8,
    "in_progress": 9, "ready": 10, "drafted": 10, "preparing": 11, "evaluated": 12,
    "evaluating": 13, "researching": 14, "lead": 15,
}

# --- 三档 colour scale (高/中/低 → green/amber/red) ---------------------------
LEVEL_FILL = {"高": "C6EFCE", "中": "FFEB9C", "低": "F8CBAD"}
NATURE_FILL = {"科研": "8EB4E3", "融合": "C3A6DE", "工程": "A9D08E"}  # blue / purple / green, saturated for clear contrast
BAND_FILL = "F4F8FC"   # every-other group-of-10 subtle band
HEADER_FILL = "305496"

COL_WIDTH = {
    "编号": 5, "类型": 11, "岗位性质": 9, "名称": 22, "地点": 16, "职位": 28,
    "工作内容 / 课题": 26, "状态": 11, "主动联系人": 26, "兴趣级": 8, "匹配度": 11, "申请截止": 11,
    "投递日期": 11, "开工日期": 11, "跟进日期": 11, "最近动作": 11, "进展记录": 34,
    "联系人": 22, "联系 / 套磁情况": 22, "合同": 16, "要求": 18, "联系人邮箱": 26,
    "轴 / Axis": 20, "申请方式": 26, "申请平台": 15, "原职位链接": 10, "材料 / 往来记录": 44, "其他 / 备注": 38,
}


def parse_frontmatter(text):
    """Tolerant top-level YAML scalar parser (skips nested/list/comment lines)."""
    m = re.match(r"^---\n(.*?)\n---", text, re.S)
    fm = {}
    if not m:
        return fm
    for line in m.group(1).splitlines():
        if not line or line[0] in " \t#-":  # indented / nested / list item / comment
            continue
        mm = re.match(r"^([A-Za-z_][\w]*):\s?(.*)$", line)
        if not mm:
            continue
        key, val = mm.group(1), mm.group(2)
        val = re.sub(r"\s+#.*$", "", val).strip()  # drop trailing "  # comment"
        if val in ("null", "~"):
            val = ""
        fm[key] = val
    return fm


def parse_block_a(text):
    """Parse the first '| Field | Value |' table under '## Block A'."""
    rows = {}
    m = re.search(r"^## Block A.*?$(.*?)(?=^## )", text, re.S | re.M)
    if not m:
        return rows
    for line in m.group(1).splitlines():
        if not line.startswith("|"):
            continue
        cells = [c.strip() for c in line.strip().strip("|").split("|")]
        if len(cells) == 2 and cells[0] and cells[0] != "Field":
            if not set(cells[0]) <= set("-: "):  # skip separator row
                rows.setdefault(cells[0], cells[1])
    return rows


def humanize_slug(slug):
    body = re.sub(r"^\d{4}-\d{2}-\d{2}-", "", slug)
    org, _, role = body.partition("--")
    return org.replace("-", " ").strip(), role.replace("_", " ").strip()


def classify(name, domain):
    blob = f"{name} {domain}".lower()
    if re.search(r"univ|institut|artorg|cardio|heg|an applied-sciences university|uva|amsterdam|"
                 r"research center|research institut|faculty", blob):
        return "高校 / 研究所"
    return "企业"


def _parse_date(s):
    m = re.search(r"(\d{4})-(\d{2})-(\d{2})", s or "")
    if not m:
        return None
    try:
        return datetime.date(int(m.group(1)), int(m.group(2)), int(m.group(3)))
    except ValueError:
        return None


def bulletize(text):
    """Render verbose free-text as bullets (split on sentence / ';' / ' · ' boundaries);
    strip markdown bold/code. Short single-phrase values pass through unchanged."""
    if not text:
        return ""
    t = text.replace("**", "").replace("`", "").strip()
    parts = re.split(r"(?:;\s+)|(?:\s+·\s+)|(?:(?<=[.])\s+(?=[A-Z(]))", t)
    parts = [p.strip().rstrip(".;").strip() for p in parts if p.strip()]
    if len(parts) <= 1:
        return t
    return "\n".join(f"- {p}" for p in parts)


def role_nature(archetype):
    """科研 (research) / 融合 (research-engineering blend) / 工程 (engineering) from
    the archetype code. 1 (Research Scientist) / 2 (Applied Scientist) -> 科研;
    3 (Research SW Engineer / applied-AI engineer with research bent) -> 融合;
    >=4 (pure SWE / MLE) -> 工程. Handles 'X -> Y' (takes the last number)."""
    nums = re.findall(r"\d+", archetype or "")
    if not nums:
        return ""
    a = int(nums[-1])
    return "科研" if a <= 2 else ("融合" if a == 3 else "工程")


def interest_level(priority):
    return {"A": "高", "B": "中", "C": "低"}.get((priority or "").strip().upper()[:1], "")


def fit_bucket(fit):
    """Return (level, numeric). high >=4.0, mid 3.0-3.99, low <3.0."""
    try:
        v = float(fit)
    except (TypeError, ValueError):
        return "", None
    if v >= 4.0:
        return "高", v
    if v >= 3.0:
        return "中", v
    return "低", v


def progress_bullets(fm, status):
    """Synthesize a dated bullet timeline from frontmatter (always reflects the md)."""
    closed = (status or "") in CLOSED_STATUS
    ev = []
    d = _parse_date(fm.get("source", ""))
    if d:
        ev.append((d, "发现 / 缓存 JD"))
    d = _parse_date(fm.get("materials_ready_date", ""))
    if d:
        ev.append((d, "材料就绪"))
    applied = _parse_date(fm.get("applied_date", ""))
    if applied:
        ev.append((applied, "已投递 ✅"))
    else:
        d = _parse_date(fm.get("submit_by", ""))
        if d:
            ev.append((d, "计划提交"))
    d = _parse_date(fm.get("last_action_date", ""))
    if d:
        ev.append((d, "最近动作"))
    d = _parse_date(fm.get("rejected_date", ""))
    if d:
        ev.append((d, "已拒 ❌"))
    d = _parse_date(fm.get("follow_up_after", ""))
    if d:
        due = (d <= TODAY) and not closed and applied is not None
        ev.append((d, "跟进 ⏰已到" if due else "跟进"))
    seen, out = set(), []
    for dt, lab in sorted(ev):
        if (dt, lab) in seen:
            continue
        seen.add((dt, lab))
        out.append(f"- {dt} {lab}")
    return "\n".join(out)


def discover_materials(folder):
    """One merged 材料 / 往来 cell: a bulleted inventory of every application material AND
    correspondence draft in the folder (CV, motivation / cover letter, portfolio, referees,
    outreach / reply / email / LinkedIn-message drafts, application-questions, the application.md).
    Diploma / transcript / certificate scans are NOT listed (PII, local). The cell hyperlinks to
    the folder (folder_dir), since a spreadsheet cell can hold only one hyperlink."""
    folder_dir = str(folder.relative_to(APPS_DIR)) + "/"
    PII = re.compile(r"diploma|diplome|transcript|certificate|doctorant_|_signed", re.I)
    items = []
    if (folder / "application.md").exists():
        items.append(("申请说明", "application.md"))
    for p in sorted(folder.glob("*.pdf")):
        n = p.name
        if "diff" in n.lower() or PII.search(n):
            continue
        low = n.lower()
        if re.search(r"_cv\.pdf$|cv\.pdf$|curriculum", low):   label = "CV"
        elif re.search(r"letter|motivation", low):             label = "动机信 / cover letter"
        elif "portfolio" in low:                               label = "作品集"
        elif "referee" in low or "reference" in low:           label = "推荐人"
        else:                                                  label = "材料"
        items.append((label, n))
    for p in sorted(folder.glob("*.md")):
        n = p.name
        if n == "application.md" or "diff" in n.lower():
            continue
        low = n.lower()
        if "outreach" in low:                                  label = "套磁 / outreach 草稿"
        elif "reply" in low:                                   label = "回信草稿"
        elif "question" in low or "answer" in low:             label = "申请问答"
        elif "email" in low or "message" in low:               label = "邮件 / 消息草稿"
        else:                                                  label = "笔记"
        items.append((label, n))
    # First bullet is the folder itself, written as the relative path, so the cell says where
    # the materials live and the single cell-level hyperlink has something to point at.
    materials = "\n".join([f"- 文件夹: {folder_dir}"] + [f"- {lab}: {nm}" for lab, nm in items])
    # Per-file links. An xlsx cell can hold only ONE hyperlink, so the bullets inside the
    # 材料 cell cannot each be clickable. Each of the common material types therefore gets
    # its own narrow column, and the 文件夹 column still covers everything else.
    links = {}
    for lab, nm in items:
        key = {"申请说明": "lnk_md", "CV": "lnk_cv", "动机信 / cover letter": "lnk_letter",
               "推荐人": "lnk_ref", "邮件 / 消息草稿": "lnk_mail",
               "作品集": "lnk_portfolio"}.get(lab)
        if key and key not in links:
            links[key] = folder_dir + nm
    return {"folder_dir": folder_dir, "materials": materials, **links}


def infer_platform(fm):
    """Infer the application platform / channel from the apply_route, url, and cv_used.
    LinkedIn Easy Apply is called out specially (coloured in the sheet)."""
    blob = " ".join(str(fm.get(k, "")) for k in
                    ("apply_route", "url", "cv_used", "links", "linkedin_apply_url")).lower()
    if "easy apply" in blob or "easyapply" in blob:
        return "LinkedIn EasyApply"
    checks = [
        ("join.com", ("join.com",)),
        ("Oxford CoreHR", ("corehr",)),
        ("Workday", ("myworkdayjobs", "workday")),
        ("SmartRecruiters", ("smartrecruiters",)),
        ("SUPSI form", ("supsi.ch",)),
        ("Greenhouse", ("greenhouse",)),
        ("Ashby", ("ashbyhq", "ashby")),
        ("Lever", ("lever.co", "jobs.lever")),
        ("werecruit", ("werecruit",)),
        ("BambooHR", ("bamboohr",)),
        ("Workable", ("workable",)),
        ("AcademicTransfer", ("academictransfer",)),
        ("Varbi", ("varbi",)),
        ("jobs.ac.uk", ("jobs.ac.uk",)),
        ("eRecruitment", ("erecruitment", "e-recruitment")),
        ("DataCareer", ("datacareer",)),
    ]
    for label, needles in checks:
        if any(n in blob for n in needles):
            return label
    if "linkedin.com/jobs" in blob:
        return "LinkedIn"
    if "http" in blob:
        return "company site"
    return ""


AXIS_RULES = [
    # Ordered most specific first. A role can match several; the first two are reported.
    # His core axes come before the generic ones so a graph role never reads as "generic ML".
    ("金融 / Finance",        r"financ|fintech|quant|bank|trading|payment|fraud|money.launder|alpha|invest|金融|风控"),
    ("水文环境 / Water-Env",  r"hydrolog|water|river|flood|catchment|climate|weather|meteo|earth observ|remote sens|forest|glacier|环境|水文|气候|遥感"),
    # Bare 预测 over-matched: it is the generic Chinese word for prediction, so "预测式模型"
    # (predictive models, Meridian Simulation) and "结构预测" (structure prediction, Roche) were read as
    # forecasting. Excluding those two compounds drops both false positives and keeps all 11
    # genuinely temporal roles, including 交通预测, 时空预测 and 波形预测. Measured 2026-07-27.
    ("时序预测 / Time-series", r"time.?series|forecast|nowcast|anomaly detection|时序|时间序列|(?<!结构)预测(?!式)"),
    ("图 / 时空 / Graph-ST",  r"\bgraph|gnn|geometric|spatio|spatial|mesh|topolog|network analysis|图|时空|拓扑"),
    ("能源 / Energy",         r"energy|power system|grid|battery|wind|solar|能源|电网"),
    ("智能体 / 平台 / Agentic", r"agent|llm|large language|language model|\brag\b|platform|mlops|orchestrat|智能体|平台|大模型"),
    ("化学 / 分子 / Chem",     r"chemi|molecul|drug|compound|redox|化学|分子"),
    ("生物医学 / Biomed",      r"biomedical|medical|clinical|health|cardio|onco|genom|protein|医学|健康|临床"),
    ("视觉 / Vision",         r"computer vision|segmentation|object detection|视觉|图像"),
    ("统计 / 概率 / Stats",    r"statistic|probabilistic|bayesian|uncertainty|统计|概率|贝叶斯"),
    ("基础 ML / Fundamental",  r"fundamental (machine learning|ml)|representation learning|out.of.distribution|generalis|generaliz|基础机器学习|表示学习"),
    ("教职 / Faculty",         r"professor|lecturer|tenure|faculty|assistant prof|教授|讲师"),
]


def axis_of(rec):
    """Classify a role onto the applicant's axes, from the role topic, position and contract text.
    Added 2026-07-27 at his request so the sheet can be filtered by axis. A role can touch several
    axes; the first two matches are reported, most specific rule first."""
    blob = " ".join(str(rec.get(k, "") or "") for k in
                    ("position", "role", "name", "requirements", "contract", "notes", "role_nature")).lower()
    hits = [label for label, pat in AXIS_RULES if re.search(pat, blob)]
    return " + ".join(hits[:2]) if hits else "其他 / Other"


def app_record(folder):
    text = (folder / "application.md").read_text(encoding="utf-8")
    fm = parse_frontmatter(text)
    a = parse_block_a(text)
    name, role_from_slug = humanize_slug(folder.name)
    domain = fm.get("role_topic", "") or a.get("Domain", "")
    status = fm.get("status", "")
    _email_src = " ".join([fm.get("apply_route", ""), fm.get("contact_name", ""),
                           fm.get("contact_email", ""), fm.get("links", "")])
    emails = list(dict.fromkeys(re.findall(r"[\w.+-]+@[\w.-]+\.\w{2,}", _email_src)))
    fit_level, fit_num = fit_bucket(fm.get("fit_score", ""))
    fit_display = fit_level
    mat = discover_materials(folder)
    return {
        "type": classify(name, domain),
        "role_nature": role_nature(fm.get("archetype", "")),
        "name": name,
        "location": fm.get("location_zh", "") or a.get("Remote", ""),
        "position": role_from_slug or a.get("Function", ""),
        "role": domain,
        "status": status,
        "interest": interest_level(fm.get("priority", "")),
        "fit_display": fit_display,
        "fit_num": fit_num if fit_num is not None else 0.0,
        "deadline": fm.get("deadline", ""),
        "applied_date": fm.get("applied_date", ""),
        "start_date": a.get("Start date", ""),
        "follow_up_after": fm.get("follow_up_after", ""),
        "last_action_date": fm.get("last_action_date", ""),
        "progress": progress_bullets(fm, status),
        "contact_name": fm.get("contact_name", ""),
        "outreach": bulletize(fm.get("outreach") or fm.get("salvage_status", "")),
        "contract": fm.get("contract", "") or a.get("Contract", ""),
        "requirements": bulletize(a.get("Seniority", "") or fm.get("requirements", "")),
        "contact_email": "; ".join(emails),
        "apply_route": bulletize(fm.get("apply_route", "") or fm.get("linkedin_apply_url", "")),
        "platform": infer_platform(fm),
        "url": fm.get("url", ""),
        "materials": mat["materials"],
        "folder_dir": mat["folder_dir"],
        # per-file link targets discovered above (lnk_cv, lnk_letter, ...)
        **{k: v for k, v in mat.items() if k.startswith("lnk_")},
        "archived": "_archive" in folder.parts,
        "notes": ("- 已归档\n" if "_archive" in folder.parts else "") + bulletize(fm.get("notes", "") or a.get("TL;DR", "")) + (("\n🔗 " + fm.get("links", "")) if fm.get("links") else ""),
        "_slug": folder.name,
    }


def _sort_key(rec):
    """Action view. tier 0 urgent-unsubmitted-deadline, 1 ready-to-submit,
    2 submitted/in-flight, 3 closed; then 兴趣级, 匹配度, stage, 科研-first, name."""
    s = rec.get("status", "")
    closed = s in CLOSED_STATUS or rec.get("archived", False)
    submitted = bool(rec.get("applied_date")) or s in SUBMITTED_STATUS
    dl = _parse_date(rec.get("deadline", ""))
    days = (dl - TODAY).days if dl else None
    urgent = (not submitted) and (days is not None) and (0 <= days <= 10)
    tier = 3 if closed else 0 if urgent else 1 if not submitted else 2
    interest_rank = {"高": 3, "中": 2, "低": 1}.get(rec.get("interest", ""), 0)
    research_first = 0 if rec.get("role_nature") == "科研" else 1
    return (tier, -interest_rank, -rec.get("fit_num", 0.0),
            STAGE_RANK.get(s, 7), research_first, rec.get("name", ""))


def load_apps():
    recs = []
    paths = sorted(APPS_DIR.glob("20*--*")) + sorted((APPS_DIR / "_archive").glob("20*--*"))
    for d in paths:
        if d.is_dir() and (d / "application.md").exists():
            recs.append(app_record(d))
    recs.sort(key=_sort_key)
    for i, r in enumerate(recs, 1):
        r["id"] = i
        r["staleness"] = staleness(r)
    for _r in recs:
        _r["axis"] = axis_of(_r)
    return recs


def _deadline_fill(rec):
    """Colour the 申请截止 cell by urgency (unsubmitted) / mute if past+submitted."""
    dl = _parse_date(rec.get("deadline", ""))
    if not dl:
        return None
    submitted = bool(rec.get("applied_date")) or rec.get("status") in SUBMITTED_STATUS
    days = (dl - TODAY).days
    if submitted:
        return "D9D9D9" if days < 0 else None
    if days < 0:
        return "FF7C80"   # overdue + unsubmitted = red
    if days <= 7:
        return "FFC000"   # within a week = orange
    if days <= 30:
        return "FFF2CC"   # within a month = pale yellow
    return None


def staleness(rec):
    """Say out loud when a row has gone out of date, so the Excel self-diagnoses.

    Four cases, checked in order of how much they should bother the applicant:
      1. the deadline passed and nothing was ever sent
      2. a package is sitting finished with the deadline inside a week
      3. an application has been silent for a long time with no outcome recorded
      4. nothing worth saying
    Closed rows (rejected, declined, withdrawn, ...) never nag: the decision is made.
    """
    status = (rec.get("status") or "").strip()
    if status in CLOSED_STATUS:
        return "-"

    dl = _parse_date(rec.get("deadline", ""))
    submitted = bool(rec.get("applied_date")) or status in SUBMITTED_STATUS

    if dl and not submitted:
        days = (dl - TODAY).days
        if days < 0:
            return f"过期未投 +{-days}天"
        if days <= 7:
            return f"仅剩 {days} 天"

    ad = _parse_date(rec.get("applied_date", ""))
    if ad and status == "applied":
        silent = (TODAY - ad).days
        if silent >= 60:
            return f"静默 {silent} 天，可视为无回应"
        if silent >= 30:
            return f"静默 {silent} 天"

    if status in {"ready", "drafted"} and not dl and not submitted:
        la = _parse_date(rec.get("last_action_date", ""))
        if la and (TODAY - la).days >= 21:
            return f"材料已备 {(TODAY - la).days} 天未投，需重新核实是否还开放"

    return "-"


def _staleness_fill(rec):
    v = staleness(rec)
    if v.startswith("过期未投"):
        return "FF7C80"   # red: the opportunity was lost while a package sat ready
    if v.startswith("仅剩"):
        return "FFC000"   # orange: act now
    if "可视为无回应" in v:
        return "D9D9D9"   # grey: almost certainly dead, stop counting on it
    if v.startswith("静默") or v.startswith("材料已备"):
        return "FFF2CC"   # pale yellow: worth a look
    return None


def _followup_fill(rec):
    d = _parse_date(rec.get("follow_up_after", ""))
    if not d or rec.get("status") in CLOSED_STATUS:
        return None
    return "FFD966" if d <= TODAY else None   # gold when a follow-up is due


# --------------------------------------------------------------------------- #
# Second sheet: the long-term target list (roles with no application folder yet)
# Added 2026-07-27 on the applicant's request: he wants every recently found,
# on-axis role visible in the Excel WITH ITS LINK so he can review and confirm.
# Source of truth is docs/LONG_TERM_TARGETS_*.md, parsed rather than duplicated,
# so the Excel never drifts from the document.
# --------------------------------------------------------------------------- #
LT_SHEET = "长期目标 · Long-term targets"


def _lt_doc():
    cands = sorted(REPO.glob("docs/LONG_TERM_TARGETS_*.md"))
    return cands[-1] if cands else None


def load_long_term():
    """Parse the long-term targets doc into rows. Each row keeps the axis heading it
    sat under, every markdown link found in the line, and the raw cells, so nothing
    is invented here and the Excel stays a view of the document."""
    doc = _lt_doc()
    if not doc:
        return []
    axis = ""
    rows = []
    for line in doc.read_text(errors="replace").splitlines():
        m = re.match(r"^##\s+(.*)", line)
        if m:
            axis = re.sub(r"\*+", "", m.group(1)).strip()
            continue
        if not line.startswith("|"):
            continue
        cells = [c.strip() for c in line.strip().strip("|").split("|")]
        if len(cells) < 3:
            continue
        joined = " ".join(cells)
        if set(joined.replace("|", "").strip()) <= set("-: "):
            continue
        if cells[0].lower() in ("role", "company", "#", "role | firm"):
            continue
        links = re.findall(r"\((https?://[^)\s]+)\)", line) + re.findall(r"(?<![(\w])(https?://[^)\s|]+)", line)
        seen, urls = set(), []
        for u in links:
            u = u.rstrip(".,;")
            if u not in seen:
                seen.add(u)
                urls.append(u)
        clean = [re.sub(r"\[([^\]]*)\]\([^)]*\)", r"\1", re.sub(r"\*+", "", c)).strip() for c in cells]
        rows.append({"axis": axis, "cells": clean, "urls": urls})
    return rows


def build_long_term_sheet(wb):
    rows = load_long_term()
    if not rows:
        return 0
    ws = wb.create_sheet(LT_SHEET)
    headers = ["轴 / Axis", "角色 / Role", "机构 / Org", "说明 / Notes",
               "最后确认 / Last seen", "链接 / Link", "链接 2", "链接 3"]
    ws.merge_cells("A1:H1")
    c = ws.cell(1, 1, "长期目标清单 · Long-term targets (permanent / tenure-track / 3yr+)")
    c.font = Font(bold=True, size=14, color="305496")
    ws.merge_cells("A2:H2")
    src = _lt_doc()
    c = ws.cell(2, 1, f"生成于 {TODAY} · 解析自 {src.relative_to(REPO) if src else '?'} · "
                      f"这些角色多数还没有 application 文件夹，所以不在第一个工作表里 · "
                      f"每行的链接可直接点开核对 · 「最后确认」早于一周的需要重新核验再投")
    c.font = Font(italic=True, size=9, color="808080")
    c.alignment = Alignment(wrap_text=True, vertical="top")
    hdr_fill = PatternFill("solid", fgColor=HEADER_FILL)
    for j, h in enumerate(headers, 1):
        cc = ws.cell(3, j, h)
        cc.font = Font(bold=True, color="FFFFFF", size=10)
        cc.fill = hdr_fill
        cc.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    r = 4
    for row in rows:
        cells = row["cells"]
        vals = [row["axis"], cells[0] if cells else "-", cells[1] if len(cells) > 1 else "-",
                " / ".join(cells[2:-1]) if len(cells) > 3 else (cells[2] if len(cells) > 2 else "-"),
                cells[-1] if len(cells) > 2 else "-"]
        for j, v in enumerate(vals, 1):
            cc = ws.cell(r, j, (v or "-")[:300])
            cc.alignment = Alignment(wrap_text=True, vertical="top")
            cc.font = Font(size=9)
        for k, u in enumerate(row["urls"][:3]):
            cc = ws.cell(r, 6 + k, u[:180])
            cc.hyperlink = u
            cc.font = Font(size=9, color="0563C1", underline="single")
            cc.alignment = Alignment(wrap_text=True, vertical="top")
        for j in range(1, 9):
            if ws.cell(r, j).value in (None, ""):
                ws.cell(r, j, "-").font = Font(size=9)
        r += 1
    for j, w in enumerate([20, 42, 24, 48, 26, 40, 32, 32], 1):
        ws.column_dimensions[get_column_letter(j)].width = w
    ws.row_dimensions[1].height = 22
    ws.row_dimensions[2].height = 38
    ws.freeze_panes = "C4"
    ws.auto_filter.ref = f"A3:H{r - 1}"
    return len(rows)


CT_SHEET = "联系人 · Contacts"
CONTACTS_DOC = REPO / "docs" / "CONTACTS.md"

# Suggested next step, keyed on the status of the application the contact is attached to.
# A contact is worth surfacing precisely when the role is NOT closed, or when the role is
# closed but the person is still reachable, which is the case that keeps getting forgotten.
CONTACT_ACTION = {
    "rejected":  "岗位已拒，但人还在：值得一封简短的保持联系邮件 (see docs/REJECTION_REPLIES_AUDIT_2026-07-28.md)",
    "declined":  "我方婉拒，关系可留",
    "ghosted":   "无回复：可礼貌跟进一次",
    "closed":    "岗位关闭，人可留作未来线索",
    "closed-before-application": "未投，人可留作未来线索",
    "ready":     "材料已备：投之前或投之后发一条短信息",
    "drafted":   "材料起草中：投递时一并联系",
    "applied":   "已投：可发一条简短的跟进 / 自我介绍",
    "acknowledged": "已确认收到：可发一条简短的跟进",
    "interview": "面试阶段：面试前后的沟通对象",
    "lead":      "线索阶段：套磁是这一步最有效的动作",
}


def _parse_contacts_doc():
    """Read the hand-maintained roster in docs/CONTACTS.md.

    Holds people who are not attached to any application folder, plus anyone whose value
    needs more words than a `contact_name:` field allows. Externalised into a markdown table
    on purpose, so the applicant can edit it without touching this script.
    """
    if not CONTACTS_DOC.exists():
        return []
    rows, in_table = [], False
    for line in CONTACTS_DOC.read_text(encoding="utf-8").splitlines():
        s = line.strip()
        if s.startswith("| Name "):
            in_table = True
            continue
        if in_table:
            if not s.startswith("|"):
                if rows:
                    break
                continue
            if set(s) <= set("|-: "):
                continue
            cells = [c.strip() for c in s.strip("|").split("|")]
            if len(cells) >= 5 and cells[0]:
                rows.append({"name": cells[0], "org": cells[1], "role": cells[2],
                             "rel": cells[3], "note": cells[4], "src": "docs/CONTACTS.md",
                             "status": "", "url": ""})
    return rows


def load_contacts():
    """Every named contact, from both sources, most actionable first."""
    out = list(_parse_contacts_doc())
    known = {r["name"].split("(")[0].strip().lower() for r in out}
    for rec in load_apps():
        name = str(rec.get("contact_name", "") or "").strip().strip('"').strip()
        if not name or name in {"-", "<TBD>"}:
            continue
        # Skip the ones that only record "no named contact" in one phrasing or another.
        low = name.lower()
        if any(k in low for k in ("no named", "easy apply", "无具名", "workday", "greenhouse",
                                 "招聘门户", "hiring;", "portal")) and not any(
                ch.isupper() for ch in name.split()[0][1:]):
            pass  # keep it anyway; the org is still useful, the note explains what is missing
        status = str(rec.get("status", "") or "")
        out.append({
            "name": name,
            "org": str(rec.get("company", "") or "-"),
            "role": str(rec.get("role", "") or "-"),
            "rel": str(rec.get("outreach", "") or "-"),
            "note": CONTACT_ACTION.get(status, "-"),
            "src": f"applications/{rec.get('slug', '')}/application.md",
            "status": status,
            "url": str(rec.get("url", "") or ""),
        })
    # Roster entries first (they are curated), then live roles, then closed ones.
    def rank(r):
        if r["src"].startswith("docs/"):
            return 0
        return 2 if r["status"] in CLOSED_STATUS else 1
    out.sort(key=rank)
    return out


def build_contacts_sheet(wb):
    rows = load_contacts()
    if not rows:
        return 0
    ws = wb.create_sheet(CT_SHEET)
    headers = ["姓名 / Name", "机构 / Organisation", "对方角色 / Their role",
               "关系 · 联系记录 / Relationship", "建议下一步 / Next step",
               "岗位状态 / Role status", "来源 / Source", "链接 / Link"]
    ws.merge_cells("A1:H1")
    c = ws.cell(1, 1, "联系人总表 · Contacts (含已拒岗位的人：岗位关了，人没关)")
    c.font = Font(bold=True, size=14, color="305496")
    ws.merge_cells("A2:H2")
    c = ws.cell(2, 1, f"生成于 {TODAY} · 两个来源合并：docs/CONTACTS.md 手工名册（排在最前，可自行编辑）"
                      f" + 每个 application.md 的 contact_name / outreach 字段 · "
                      f"已拒岗位的联系人依然列出，因为岗位关闭不等于关系关闭，"
                      f"这正是最容易被遗忘的一类")
    c.font = Font(italic=True, size=9, color="808080")
    c.alignment = Alignment(wrap_text=True, vertical="top")
    hdr_fill = PatternFill("solid", fgColor=HEADER_FILL)
    for j, h in enumerate(headers, 1):
        cc = ws.cell(3, j, h)
        cc.font = Font(bold=True, color="FFFFFF", size=10)
        cc.fill = hdr_fill
        cc.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    r = 4
    for row in rows:
        vals = [row["name"], row["org"], row["role"], row["rel"], row["note"],
                STATUS_ZH.get(row["status"], row["status"] or "-"), row["src"]]
        for j, v in enumerate(vals, 1):
            cc = ws.cell(r, j, (str(v) or "-")[:600])
            cc.alignment = Alignment(wrap_text=True, vertical="top")
            cc.font = Font(size=9)
        if row["src"].startswith("docs/"):
            for j in range(1, 9):
                ws.cell(r, j).fill = PatternFill("solid", fgColor="FFF2CC")
        if row["status"] in CLOSED_STATUS:
            ws.cell(r, 6).font = Font(size=9, color="C00000")
        if row["url"]:
            cc = ws.cell(r, 8, row["url"][:180])
            cc.hyperlink = row["url"]
            cc.font = Font(size=9, color="0563C1", underline="single")
            cc.alignment = Alignment(wrap_text=True, vertical="top")
        for j in range(1, 9):
            if ws.cell(r, j).value in (None, ""):
                ws.cell(r, j, "-").font = Font(size=9)
        lines = max(len(str(ws.cell(r, j).value)) // 46 + 1 for j in (4, 5))
        ws.row_dimensions[r].height = min(320, lines * 15 + 6)
        r += 1
    for j, w in enumerate([26, 24, 30, 52, 52, 16, 40, 34], 1):
        ws.column_dimensions[get_column_letter(j)].width = w
    ws.row_dimensions[1].height = 22
    ws.row_dimensions[2].height = 40
    ws.freeze_panes = "C4"
    ws.auto_filter.ref = f"A3:H{r - 1}"
    return len(rows)


def _materials_rich(text):
    """Materials cell as rich text: the first bullet (the folder path) in link blue, the rest plain.

    An xlsx cell holds at most one hyperlink, so the per-file bullets cannot each be
    clickable. Colouring the whole cell as a link made its four bullets read as one merged
    block, so only the folder line is styled.
    """
    text = str(text or "")
    if not text:
        return ""
    head, sep, tail = text.partition("\n")
    try:
        from openpyxl.cell.rich_text import CellRichText, TextBlock
        from openpyxl.cell.text import InlineFont
    except ImportError:                       # older openpyxl: fall back to plain text
        return text
    blue = InlineFont(color="0563C1", u="single", sz=9)
    black = InlineFont(color="000000", sz=10)
    if not sep:
        return CellRichText(TextBlock(blue, head))
    return CellRichText(TextBlock(blue, head), TextBlock(black, sep + tail))


# --- 主动联系人 column ------------------------------------------------------
# Added 2026-07-29 at his request: one glance at whether a role has a human route
# in, and whether that route has been used. Colour encodes urgency, not identity.
CONTACT_FILL = {
    "unused":  "FF9999",   # red    - a named human, never contacted, and the route is still open
    "pending": "FFC000",   # orange - contacted, but the role still waits on him
    "live":    "C6EFCE",   # green  - contacted and the application is in flight
    "none":    "D9D9D9",   # grey   - no named human to write to
}
# contact_name values that record the ABSENCE of a person rather than a person
_NO_PERSON = ("easy apply", "no named", "无具名", "招聘门户", "workday", "greenhouse",
              "portal", "recruitment", "hiring;", "(no name")


def _roster_companies():
    """Companies named in docs/CONTACTS.md, the hand-kept roster."""
    doc = REPO / "docs" / "CONTACTS.md"
    if not doc.exists():
        return set()
    out = set()
    for line in doc.read_text(encoding="utf-8").splitlines():
        if line.startswith("|") and line.count("|") >= 5:
            cells = [c.strip() for c in line.strip("|").split("|")]
            if len(cells) >= 2 and cells[1] and cells[1] != "Organisation":
                out.add(cells[1].split(",")[0].strip().lower())
    return out


def contact_flag(rec, roster):
    """(display text, fill key) for the 主动联系人 column."""
    name = str(rec.get("contact_name", "") or "").strip().strip('"').strip()
    low = name.lower()
    has_person = bool(name) and name not in ("-", "<TBD>", "null") and not any(k in low for k in _NO_PERSON)
    outreach = str(rec.get("outreach", "") or "").strip()
    ol = outreach.lower()
    # An outreach field can RECORD that nobody was contacted. Cortex Systems's says
    # "no direct contact"; the Meridian Simulation one says research only, nothing sent. Treating
    # any non-empty value as contact would hide exactly the routes worth using.
    _denials = ("no direct contact", "nothing was sent", "research only", "no contact",
                "未发送", "没有联系", "未联系")
    contacted = (bool(outreach) and outreach not in ("-", "null")
                 and not any(d in ol for d in _denials))
    status = str(rec.get("status", "") or "")
    # The record carries the org in `name` (e.g. "Cortex Systems Lausanne"); there is no
    # `company` key. Matching on the wrong key silently disabled the roster override.
    org = str(rec.get("name", "") or "").lower()
    in_roster = any(rc and rc != "---" and rc.split()[0] in org for rc in roster)

    if not has_person and not in_roster:
        return ("-", "none")

    short = name.split(";")[0].split("(")[0].strip()[:40] or "见 联系人 表"
    if in_roster and not has_person:
        short = "见 联系人 表"

    # A closed role can still hold the best route in. The Cortex Systems referral is on a
    # REJECTED application, and greying it out by status would hide the single highest
    # value contact in the workspace. So the roster overrides status.
    if in_roster and not contacted:
        return (short + "  ⚑", "unused")
    if status in SUBMITTED_STATUS:
        return (short, "live" if contacted else "pending")
    if status in CLOSED_STATUS:
        return (short, "none" if not in_roster else "unused")
    return (short, "pending" if contacted else "unused")


def _wrapped_lines(text, col_width):
    """How many lines Excel will draw for this cell, counting wrapping as well as newlines."""
    text = str(text or "")
    if not text:
        return 1
    usable = max(col_width - 1.5, 6)          # padding eats roughly a character and a half
    total = 0
    for logical in text.split("\n"):
        total += max(1, -(-_disp_width(logical) // usable))
    return int(total)


def build():
    recs = load_apps()
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = SHEET
    ncol = len(COLUMNS)
    last_col = get_column_letter(ncol)
    thin = Side(style="thin", color="D9D9D9")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    ws.merge_cells(f"A1:{last_col}1")
    c = ws.cell(1, 1, "申请进度总表 · Applications Tracker")
    c.font = Font(bold=True, size=14, color="305496")

    ws.merge_cells(f"A2:{last_col}2")
    c = ws.cell(2, 1, f"生成于 {TODAY} · 由 scripts/applications_tracker.py 从各 application.md 生成 · "
                      f"排序：①临近截止未投 ②已就绪未投(按兴趣/匹配) ③已投/进行中 ④已关闭 · "
                      f"配色：兴趣级/匹配度 高=绿 中=黄 低=红；截止 红=逾期/橙≤7天/黄≤30天；跟进 金=已到 · "
                      f"点「原职位链接」看 JD、「申请材料」进 application.md · "
                      f"主动联系人：红=还有没用过的人脉动作(最该动的，⚑=名册 docs/CONTACTS.md 里有此人) 橙=已联系、岗位待你行动 绿=已联系且在流程中 灰=无具名联系人")
    c.font = Font(italic=True, size=9, color="808080")
    c.alignment = Alignment(wrap_text=True, vertical="top")

    hdr_fill = PatternFill("solid", fgColor=HEADER_FILL)
    for j, (h, _) in enumerate(COLUMNS, 1):
        c = ws.cell(3, j, h)
        c.font = Font(bold=True, color="FFFFFF", size=10)
        c.fill = hdr_fill
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = border

    _ROSTER = _roster_companies()
    for ri, rec in enumerate(recs, start=4):
        band = ((ri - 4) // 10) % 2          # alternate shade every 10 items
        base_fill = BAND_FILL if band else None
        _st = rec.get("status") or ""
        if _st == "rejected":
            base_fill = "FFD6D6"             # rejected jobs: whole row tinted red (status cell = big red)
        elif _st in CLOSED_STATUS or rec.get("archived"):
            base_fill = "E2E2E2"             # closed / withdrawn / declined / archived: whole row greyed
        # row height ~ tallest wrapping cell (progress bullets / notes / role)
        lines = max(rec.get("progress", "").count("\n") + 1,
                    rec.get("materials", "").count("\n") + 1,
                    rec.get("outreach", "").count("\n") + 1,
                    -(-len(rec.get("notes", "")) // 40),
                    -(-len(rec.get("role", "")) // 26), 1)
        ws.row_dimensions[ri].height = min(320, lines * 15 + 6)

        for j, (h, key) in enumerate(COLUMNS, 1):
            val = rec.get(key, "") or ""
            c = ws.cell(ri, j, val)
            c.alignment = Alignment(vertical="top", wrap_text=True)
            c.font = Font(size=10)
            c.border = border
            fill = base_fill
            if key in LINK_LABELS and val:
                c.value, c.hyperlink = LINK_LABELS[key], val
                c.font = Font(size=10, color="0563C1", underline="single")
                c.alignment = Alignment(horizontal="center", vertical="center")
            elif key == "materials":
                # The cell carries ONE hyperlink, to the folder, because xlsx allows no more
                # than one per cell. Rich text paints only the first bullet, the folder path,
                # in link blue; every other bullet stays plain black so the items still read
                # as separate lines. Styling the whole cell as a link was the 2026-07-29 fault.
                c.value = _materials_rich(val)
                if rec.get("folder_dir"):
                    c.hyperlink = rec["folder_dir"]
            elif key == "contact_flag":
                # Restored 2026-07-31. This branch was swallowed on 2026-07-29 when the lnk_*
                # columns were removed by replacing the whole span from the lnk_ branch to the
                # status branch; contact_flag sat between them. The column stayed in COLUMNS
                # and the classifier stayed defined, so the build kept reporting success while
                # every cell rendered as a dash.
                txt, kind = contact_flag(rec, _ROSTER)
                c.value = txt
                fill = CONTACT_FILL.get(kind, base_fill)
                if kind == "unused":
                    c.font = Font(size=9, bold=True, color="9C0006")
                else:
                    c.font = Font(size=9)
            elif key == "status":
                c.value = STATUS_ZH.get(val, val)
                fill = STATUS_FILL.get(val, base_fill)
                if val == "rejected":
                    c.font = Font(size=11, bold=True, color="FFFFFF")  # 大红 + white bold
            elif key in ("interest", "fit_display"):
                fill = LEVEL_FILL.get(val, base_fill)
                if val:
                    c.alignment = Alignment(horizontal="center", vertical="center")
            elif key == "role_nature":
                fill = NATURE_FILL.get(val, base_fill)
                c.alignment = Alignment(horizontal="center", vertical="center")
            elif key == "deadline":
                fill = _deadline_fill(rec) or base_fill
            elif key == "staleness":
                fill = _staleness_fill(rec) or base_fill
            elif key == "follow_up_after":
                fill = _followup_fill(rec) or base_fill
            elif key == "id":
                c.alignment = Alignment(horizontal="center", vertical="center")
            # force-fill: never leave a blank cell. Empty -> a muted, centred "-".
            if c.value in (None, ""):
                c.value = "-"
                c.font = Font(size=10, color="BFBFBF")
                c.alignment = Alignment(horizontal="center", vertical="center")
            if fill:
                c.fill = PatternFill("solid", fgColor=fill)

    for j, (h, _) in enumerate(COLUMNS, 1):
        ws.column_dimensions[get_column_letter(j)].width = COL_WIDTH.get(h, 15)
    ws.row_dimensions[1].height = 22
    ws.row_dimensions[2].height = 42
    ws.freeze_panes = "E4"        # keep 编号/类型/岗位性质/名称 visible
    if recs:
        ws.auto_filter.ref = f"A3:{last_col}{3 + len(recs)}"

    n_lt = build_long_term_sheet(wb)
    n_ct = build_contacts_sheet(wb)

    XLSX.parent.mkdir(parents=True, exist_ok=True)
    backup_existing_tracker()
    wb.save(XLSX)
    print(f"[build] wrote {XLSX.relative_to(REPO)}  ({len(recs)} applications, {ncol} columns)"
          f" + long-term sheet ({n_lt} targets) + contacts sheet ({n_ct} people)")


def backup_existing_tracker(keep_recent=20, keep_daily_days=60):
    """Copy the current tracker xlsx to a timestamped local backup BEFORE it is
    overwritten, then prune. Backups live under outputs/tracker_backups/ (gitignored).
    git history already versions the committed xlsx; this guards the working copy
    against a bad or partial rebuild between commits, and keeps periodic restore points.
    Retention: the most recent `keep_recent` backups, plus the newest backup per
    calendar day for the last `keep_daily_days` days."""
    if not XLSX.exists():
        return None
    BACKUP_DIR.mkdir(parents=True, exist_ok=True)
    stamp = datetime.datetime.now().strftime("%Y-%m-%d_%H%M%S")
    dest = BACKUP_DIR / f"APPLICATIONS_TRACKER.{stamp}.xlsx"
    n = 1
    while dest.exists():            # two rebuilds within the same second
        dest = BACKUP_DIR / f"APPLICATIONS_TRACKER.{stamp}_{n}.xlsx"
        n += 1
    shutil.copy2(XLSX, dest)
    _prune_backups(keep_recent, keep_daily_days)
    print(f"[backup] saved {dest.relative_to(REPO)}")
    return dest


def _backup_day(path):
    """Parse the YYYY-MM-DD date from APPLICATIONS_TRACKER.<YYYY-MM-DD>_<HHMMSS>.xlsx."""
    m = re.search(r"\.(\d{4})-(\d{2})-(\d{2})_", path.name)
    if not m:
        return None
    try:
        return datetime.date(int(m.group(1)), int(m.group(2)), int(m.group(3)))
    except ValueError:
        return None


def _prune_backups(keep_recent, keep_daily_days):
    """Keep the most recent `keep_recent` backups plus the newest per day within
    `keep_daily_days`; delete the rest. Bounds disk while keeping ~2 months of
    daily restore points."""
    backups = sorted(BACKUP_DIR.glob("APPLICATIONS_TRACKER.*.xlsx"))
    if len(backups) <= keep_recent:
        return
    keep = set(backups[-keep_recent:])
    newest_per_day = {}
    for f in backups:
        day = _backup_day(f)
        if day is None:
            keep.add(f)            # unparseable name -> keep, do not risk deleting
            continue
        newest_per_day[day] = f    # ascending sort -> last write per day wins
    for day, f in newest_per_day.items():
        if (TODAY - day).days <= keep_daily_days:
            keep.add(f)
    for f in backups:
        if f not in keep:
            try:
                f.unlink()
            except OSError:
                pass


def cmd_list():
    for r in load_apps():
        print(f"{r['id']:>2}  [{r['status']:<11}] {r['type']:<10} {r['role_nature']:<4} "
              f"{r['name'][:26]:<26} 兴趣={r['interest'] or '-'} 匹配={r['fit_display'] or '-':<9} "
              f"applied={r['applied_date'] or '-':<11} 截止={r['deadline'] or '-'}")


def _set_md_field(md_path, key, value):
    """Update one top-level frontmatter scalar, preserving a trailing '# comment'. Returns (changed, old)."""
    text = md_path.read_text(encoding="utf-8")
    pat = re.compile(rf"^({re.escape(key)}:)[ \t]*(.*?)([ \t]+#.*)?$", re.M)
    m = pat.search(text)
    newval = value if value not in ("", None) else "null"
    if not m:
        return (False, None)
    old = m.group(2).strip()
    if old == str(newval).strip():
        return (False, old)
    new_line = f"{m.group(1)} {newval}{m.group(3) or ''}"
    md_path.write_text(text[:m.start()] + new_line + text[m.end():], encoding="utf-8")
    return (True, old)


def cmd_to_md(slug=None):
    if not XLSX.exists():
        sys.exit("xlsx not found — run `build` first.")
    wb = openpyxl.load_workbook(XLSX)
    ws = wb[SHEET]
    keys = [k for _, k in COLUMNS]
    link_col = keys.index("folder_link") + 1
    colidx = {k: i + 1 for i, k in enumerate(keys)}
    n = 0
    for ri in range(4, ws.max_row + 1):
        link = ws.cell(ri, link_col).hyperlink
        target = link.target if link else None
        if not target:
            continue
        rel = target.split("/application.md")[0]
        if slug and rel != slug:
            continue
        md = APPS_DIR / rel / "application.md"
        if not md.exists():
            continue
        for f in TRACK_TO_MD:
            val = ws.cell(ri, colidx[f]).value
            if val is None:
                continue
            if f == "status":
                val = STATUS_ZH_REV.get(str(val).strip(), val)
            changed, old = _set_md_field(md, f, str(val))
            if changed:
                print(f"[to-md] {rel}: {f}: {old!r} -> {val!r}")
                n += 1
    print(f"[to-md] {n} field(s) updated.")


def cmd_set(slug, key, value):
    md = APPS_DIR / slug / "application.md"
    if not md.exists():
        sys.exit(f"not found: {md}")
    changed, old = _set_md_field(md, key, value)
    print(f"[set] {slug}: {key}: {old!r} -> {value!r}" if changed else f"[set] no change ({key} already {old!r} or absent)")
    build()


def main():
    cmd = sys.argv[1] if len(sys.argv) > 1 else "build"
    if cmd == "build":
        build()
    elif cmd == "list":
        cmd_list()
    elif cmd == "to-md":
        cmd_to_md(sys.argv[2] if len(sys.argv) > 2 else None)
    elif cmd == "set":
        if len(sys.argv) < 5:
            sys.exit("usage: set SLUG KEY VALUE")
        cmd_set(sys.argv[2], sys.argv[3], " ".join(sys.argv[4:]))
    else:
        sys.exit(__doc__)


if __name__ == "__main__":
    main()
